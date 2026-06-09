# 提单台账生成器（通用·零 DOM 依赖）：读 issues.json + 模板 config → Excel + 嵌截图。
# 列定义不写死，从 template-config.json 的 columns 读——换模板=换 config，台账列跟着变。
#
# 用法：python3 gen_xlsx.py [项目目录] [--config=模板config路径]
#   项目目录    读 {目录}/issues.json（无则 issues.example.json）、截图 {目录}/shots/、可选 {目录}/notes.md
#   --config    模板 config 路径；省略=用 templates/acceptance-templates/default/template-config.json
#   输出        {目录}/问题清单.xlsx
#
# issues.json 每条按模板 columns 的 key 填字段；judging/处理状态列由模板 fillRule 规定"留空给人"，AI 不自动填。
import json, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from PIL import Image as PILImage

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(SCRIPT_DIR, "..", "..", "..", ".."))  # scripts→skill→skills→.claude→repo

args = [a for a in sys.argv[1:] if not a.startswith("--")]
opts = dict(a[2:].split("=", 1) for a in sys.argv[1:] if a.startswith("--") and "=" in a)
BASE = args[0] if args else os.getcwd()
CONFIG = opts.get("config") or os.path.join(REPO, "templates", "acceptance-templates", "default", "template-config.json")

with open(CONFIG, encoding="utf-8") as f:
    cfg = json.load(f)
columns = cfg["columns"]
render = cfg.get("render", {})
CN = render.get("font", "PingFang SC")
HEADER_FILL = render.get("headerFill", "4A7DD4")
THUMB_W = render.get("thumbWidth", 400)

ISSUES = os.path.join(BASE, "issues.json")
if not os.path.exists(ISSUES):
    ISSUES = os.path.join(BASE, "issues.example.json")
SHOTS = os.path.join(BASE, "shots")
NOTES = os.path.join(BASE, "notes.md")
OUT = os.path.join(BASE, "问题清单.xlsx")
with open(ISSUES, encoding="utf-8") as f:
    issues = json.load(f)

wb = Workbook()
ws = wb.active
ws.title = "问题清单"
thin = Side(style="thin", color="D0D7DE")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
head_fill = PatternFill("solid", fgColor=HEADER_FILL)
head_font = Font(name=CN, size=11, bold=True, color="FFFFFF")
cell_font = Font(name=CN, size=10, color="1F2328")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 表头（从 columns 读）
for c, col in enumerate(columns, 1):
    cell = ws.cell(1, c, col.get("title", col["key"]))
    cell.fill = head_fill; cell.font = head_font; cell.alignment = center; cell.border = border
    w = col.get("width", 16)
    if col.get("type") == "image":
        w = max(w, THUMB_W / 7.0 + 2)  # 图列要放得下缩略图，免得图溢到旁边列
    ws.column_dimensions[chr(64 + c)].width = w
ws.row_dimensions[1].height = 26
ws.freeze_panes = "A2"

os.makedirs(SHOTS, exist_ok=True)
thumbs = os.path.join(SHOTS, "thumbs"); os.makedirs(thumbs, exist_ok=True)
img_keys = {col["key"] for col in columns if col.get("type") == "image"}

for r, it in enumerate(issues, start=2):
    rowh = 60
    for c, col in enumerate(columns, 1):
        key = col["key"]
        align = Alignment(horizontal=col.get("align", "center"), vertical="center", wrap_text=col.get("wrap", True))
        if key in img_keys:
            shot = str(it.get(key, ""))
            if shot.endswith(".png") and os.path.exists(os.path.join(SHOTS, shot)):
                thumb = os.path.join(thumbs, "thumb_" + shot)
                im = PILImage.open(os.path.join(SHOTS, shot)); im.thumbnail((THUMB_W, THUMB_W), PILImage.LANCZOS); im.save(thumb)
                xim = XLImage(thumb)
                # TwoCellAnchor：图卡进 D{r} 这一个单元格、随单元格走（不再浮在网格上方乱跑）
                xim.anchor = TwoCellAnchor(
                    editAs="twoCell",
                    _from=AnchorMarker(col=c - 1, row=r - 1, colOff=0, rowOff=0),
                    to=AnchorMarker(col=c, row=r, colOff=0, rowOff=0),
                )
                ws.add_image(xim)
                rowh = max(rowh, im.size[1] * 0.75 + 4)  # 行高贴合图高、配合列宽让单元格≈图比例、twoCell 不变形
                cell = ws.cell(r, c, "")
            else:
                cell = ws.cell(r, c, shot or "—")
        else:
            cell = ws.cell(r, c, it.get(key, ""))
        cell.font = cell_font; cell.border = border; cell.alignment = align
    ws.row_dimensions[r].height = rowh

# 说明页：读 notes.md（行首 # 加粗），无则放写法提示
ws2 = wb.create_sheet("说明")
ws2.column_dimensions["A"].width = 100
if os.path.exists(NOTES):
    for r, line in enumerate(open(NOTES, encoding="utf-8").read().splitlines(), 1):
        cell = ws2.cell(r, 1, line.lstrip("#").strip())
        cell.font = Font(name=CN, size=11, bold=line.lstrip().startswith("#"))
        cell.alignment = Alignment(vertical="center", wrap_text=True)
else:
    tip = ["【说明页】技术根因 / 根因簇(A/B/C…标先修哪个) / 已复测项 / 待确认项 / 测试账号授权 写进同目录 notes.md（行首 # 加粗），重跑即生成。"]
    for r, t in enumerate(tip, 1):
        ws2.cell(r, 1, t).font = Font(name=CN, size=11)

wb.save(OUT)
print("SAVED:", OUT)
print("模板:", cfg.get("templateInfo", {}).get("name", "?"), "| 列:", [c["key"] for c in columns], "| rows:", len(issues))
